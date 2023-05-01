import pandas as pd 
import os
#from pprint import pprint
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
#from geopy.geocoders import Nominatim
#import geocoder
import datetime

country_currency = {
    'United States of America': 'USD',
    'United Kingdom': 'GBP',
    'Canada': 'CAD',
    'Australia': 'AUD',
    'Germany': 'EUR',
    'France': 'EUR',
    'Italy': 'EUR',
    'Spain': 'EUR',
    'Netherlands': 'EUR',
    'Belgium': 'EUR',
    'Switzerland': 'CHF',
    'Sweden': 'SEK',
    'Norway': 'NOK',
    'Denmark': 'DKK',
    'Japan': 'JPY',
    'China': 'CNY',
    'India': 'INR',
    'Brazil': 'BRL',
    'Mexico': 'MXN',
    'Argentina': 'ARS',
    'Russia': 'RUB',
    'South Africa': 'ZAR',
    'Saudi Arabia': 'SAR',
    'United Arab Emirates': 'AED',
    'Turkey': 'TRY',
    'South Korea': 'KRW',
    'Singapore': 'SGD',
    'Hong Kong': 'HKD',
    'New Zealand': 'NZD',
    'Poland': 'PLN',
    'Indonesia': 'IDR',
    'Malaysia': 'MYR',
    'Thailand': 'THB',
    'Philippines': 'PHP',
    'Egypt': 'EGP',
    'Israel': 'ILS',
    'Colombia': 'COP',
    'Chile': 'CLP',
    'Peru': 'PEN',
    'Nigeria': 'NGN',
    'Kenya': 'KES',
    'Morocco': 'MAD',
    'Pakistan': 'PKR',
    'Vietnam': 'VND',
    'Ukraine': 'UAH',
    'Czech Republic': 'CZK',
    'Czechia': 'CZK',
    'Romania': 'RON',
    'Hungary': 'HUF',
    'Portugal': 'EUR',
    'Greece': 'EUR',
    'Ireland': 'EUR',
    'Austria': 'EUR',
    'Finland': 'EUR',
    'Bulgaria': 'BGN',
    'Croatia': 'HRK',
    'Slovakia': 'EUR',
    'Slovenia': 'EUR',
    'Lithuania': 'EUR',
    'Latvia': 'EUR',
    'Estonia': 'EUR',
    'Iceland': 'ISK',
    'Luxembourg': 'EUR',
    'Malta': 'EUR',
    'Cyprus': 'EUR',
    'Albania': 'ALL',
    'Bosnia and Herzegovina': 'BAM',
    'Macedonia': 'MKD',
    'Montenegro': 'EUR',
    'Serbia': 'RSD',
    'Belarus': 'BYN',
    'Kazakhstan': 'KZT',
    'Moldova': 'MDL',
    'Georgia': 'GEL',
    'Armenia': 'AMD',
    'Qatar': 'QAR',
    'Kuwait': 'KWD',
    'Oman': 'OMR',
    'Bahrain': 'BHD',
    'Iraq': 'IQD',
    'Iran': 'IRR',
    'Lebanon': 'LBP',
    'Jordan': 'JOD',
    'Syria': 'SYP',
    'Yemen': 'YER',
    'Afghanistan': 'AFN',
    'Uzbekistan': 'UZS'}


excel_file = 'Customer-Onboarding-Agent.xlsx'
wb = load_workbook(excel_file, data_only = True)
sh = wb['Entities']



#pd.set_option('display.max_rows', 500)
#pd.set_option('display.max_columns', 500)
#pd.set_option('display.width', 1000)
#df=pd.read_csv('customer_data.csv')
df_xlsx=pd.ExcelFile('Customer-Onboarding-Agent.xlsx')
df1 = pd.read_excel(df_xlsx, 'Entities')
df2 = pd.read_excel(df_xlsx, 'Format Templates')

#print(df1[['Company name', 'Registrar', 'Company number']])

#print(df1.iloc[2:5]) # get a specific row amount (2-5 range)

#print(df1.iloc[3,3]) # this is anankes comopany number get a specific box value (row 2, column 4 ) [starts from 0 count below headers]


#for index, row in df1.iterrows():  
 #   print(index, row['Company name']) # get the company names and their index, can loop throu all items
#
#print (df1.loc[(df1['Registered Country'] == "United States of America") | (df1["Entity group"]=="AMER")])#look for specific values in columns and return those


#pprint(df1.sort_values("Company number", ascending=False))
# df1['Test'] = df1['VAT number'] + df1['Entity group'] #inserting new header and column data

#df1['Test'] = df1.iloc[3,2]
#print(df1.head(5))


cols = list(df1.columns.values)
print (cols)
#df1=df1[cols[0:4] + [cols[-1]]+cols[4:8]] reconstructing the table with only my wanted cols 
sorted_df = df1.sort_values(by='Company number')

#pprint(df1.sort_values("Company number", ascending=False))


# Iterate through the sorted column and check for duplicates
column_name = 'Company number'
prev_value = None
for index, row in sorted_df.iterrows():
    for column_index, column_value in enumerate(row):
        if sorted_df.columns[column_index] == column_name:
            current_value = column_value
            print("Row index:", index)
            print("Column index:", column_index)
            print("Column value:", current_value)

            if current_value == prev_value:
                print("BREAKPOINt")
                sh.cell(row=int(index)+2, column=int(column_index)+1).comment = Comment("DUPLICATE FOUND!", "DARIUS")
                sh.cell(row=int(prev_index)+2, column=int(prev_column)+1).comment = Comment("DUPLICATE FOUND!", "DARIUS")

                #sh.cell(row=int()+1, column=int(column_index)+1).comment = Comment("DUPLICATE FOUND!", "DARIUS")
                #sh.cell(row=int(index)+2, column=int(column_index)+1).comment = Comment("DUPLICATE FOUND!", "DARIUS")

                print("Duplicate found:", current_value)
            if len(str(current_value))>8:
                sh.cell(row=int(index)+2, column=int(column_index)+1).comment = Comment("COMPANY NUMBER TYPO FOUND, IS THIS MEANT TO BE BIGGER THAN 8 CHARACTERS?", "DARIUS")
            prev_index=index
            prev_column=column_index
            prev_value = current_value

column_name = 'Registered inspection address'            
for index, row in sorted_df.iterrows():
    for column_index, column_value in enumerate(row):
        if sorted_df.columns[column_index] == column_name:
            current_value = column_value
            print("Row index:", index)
            print("Column index:", column_index)
            print("Column value:", current_value)
            if pd.isnull(current_value):
                sh.cell(row=int(index)+2, column=int(column_index)+1).comment = Comment("INSPECTION ADDRESS NOT INPUTTED!", "DARIUS")

# Get the column from the Pandas DataFrame
column_name = 'Registered address'

column_values = sorted_df[column_name].str.replace(",", "\n")

# Overwrite the rows with the column values
for index, row in sorted_df.iterrows():
    current_value = column_values[index]
    print(current_value)
    print("Row index:", index)
    print("Column value:", current_value)

    if pd.isnull(current_value):
        print("Cell value is empty")
        sh.cell(row=int(index)+2, column=int(column_index)+1).comment = Comment("REGISTERED ADDRESS NOT INPUTTED!", "DARIUS")
    else:
        cell = sh.cell(row=index+2, column=5)
        cell.value = current_value
        cell.alignment = Alignment(wrap_text=True)
        if current_value.count("\n")!=6:
            comment = Comment(f"Standard Not followed", "Darius")
            cell.comment = comment

column_name = 'Functional currency'
for index, row in sorted_df.iterrows():
    for column_index, column_value in enumerate(row):
        if df1.columns[column_index] == column_name:
            currency=(column_value[0:3])
            country=df1.iloc[index,4][df1.iloc[index,4].rfind(",")+2:]
            if country in country_currency and country_currency[country] == currency:
                print(f"Match found! Country: {country}, Currency: {currency}")
            else:
                print(f"No match found! Country: {country}, Currency: {currency}")
                sh.cell(row=int(index)+2, column=int(column_index)+1).comment = Comment("MISMATCH CURRENCY CODE", "DARIUS")


column_name = 'Incorporation date'
for index, row in sorted_df.iterrows():
    for column_index, column_value in enumerate(row):
        if df1.columns[column_index] == column_name:
            date=column_value
            if date > datetime.date.today():
                sh.cell(row=int(index)+2, column=int(column_index)).comment = Comment("Incorporation Date WRONG! Check date", "DARIUS")


column_name="Risk factors"
column_values = sorted_df[column_name]

for index, row in sorted_df.iterrows():
    current_value = column_values[index]
    print(current_value)
    print("Row index:", index)
    print("Column value:", current_value)

    if pd.isnull(current_value):
        print("Cell value is empty")
        sh.cell(row=int(index)+2, column=int(column_index)).comment = Comment("Risk factors missing", "DARIUS")

wb.save('commented_sample.xlsx')

"""ADDRESS VALIDATION / SANITASIATION
geolocator = Nominatim(user_agent='ThisIsKubernoAgentAtWork')

for index, row in df1.iterrows():
    address = row['Registered address']
    print(address)
    functional_currency = row['Functional currency']

    # Extract country from the address using geopy
    location = geolocator.geocode(address)
    glocation=geocoder.google(address)
    print(location,glocation)
    if location:
        country = location.raw['address'].get('country')
        if country in country_currency:
            print ( country,country_currency)
            expected_currency = country_currency[country]
            if functional_currency == expected_currency:
                # Functional currency is correct for the country
                # Add your logic here
                pass
            else:
                print ("incorrect entry", country,country_currency)

                # Functional currency is incorrect for the country
                # Add your logic here
                pass
        else:
            # Country is not found in the dictionary
            # Add your logic here
            pass
    else:
        # Failed to geocode address
        # Add your logic here
        pass"""
"""
for index, row in df1.iterrows():
    country = row['Registered address']
    functional_currency = row['Functional currency']
    print(country)
    if functional_currency.count("")!=6:
    if country in country_currency:
        print(functional_currency,country)
        expected_currency = country_currency[country]
        if functional_currency == expected_currency:
            print(functional_currency,country)
            # Functional currency is correct for the country
            # Add your logic here
            pass
        else:
            print("missmatch found", expected_currency,functional_currency,country)
            # Functional currency is incorrect for the country
            # Add your logic here
            pass
    else:
        print("not found", functional_currency,country)
        # Country is not found in the dictionary
        # Add your logic here
        pass
"""

"""for index, row in sorted_df.iterrows():
    for column_index, column_value in enumerate(row):
        if sorted_df.columns[column_index] == column_name:
            currency=column_value

            if column_value in country_currency:
                expected_currency = country_currency[column_value]
                if functional_currency == expected_currency:"""


"""for index, row in df1.iterrows():
    country = row['Registered inspection address']
    functional_currency = row['Functional currency']

    if country in country_currency:
        expected_currency = country_currency[country]
        if expected_currency in functional_currency :
            # Functional currency is correct for the country
            # Add your logic here
            pass
        else:
            sh.cell(row=int(index)+2, column=int(column_index)+1).comment = Comment("CURRENCY IS INCORRECT", "DARIUS")
            # Functional currency is incorrect for the country
            # Add your logic here
            pass
    else:
        # Country is not found in the dictionary
        # Add your logic here
        pass"""

"""
for index, row in sorted_df.iterrows():
    for column_index, column_value in enumerate(row):
        if sorted_df.columns[column_index] == column_name:
            current_value = column_value
            print("Row index:", index)
            print("Column index:", column_index)
            print("Column value:", current_value)
"""

"""for index, row in sorted_df.iterrows():
    for column_index, column_value in enumerate(row):
        if sorted_df.columns[column_index] == column_name:
            current_value = column_value
            print("Row index:", index)
            print("Column index:", column_index)
            print("Column value:", current_value)

            if current_value == prev_value:
                print("BREAKPOINt")
                sh.cell(row=int(index)+2, column=int(column_index)+1).comment = Comment("DUPLICATE FOUND!", "DARIUS")
                sh.cell(row=int(prev_index)+2, column=int(prev_column)+1).comment = Comment("DUPLICATE FOUND!", "DARIUS")

                #sh.cell(row=int()+1, column=int(column_index)+1).comment = Comment("DUPLICATE FOUND!", "DARIUS")
                #sh.cell(row=int(index)+2, column=int(column_index)+1).comment = Comment("DUPLICATE FOUND!", "DARIUS")

                print("Duplicate found:", current_value)
            prev_index=index
            prev_column=column_index
            prev_value = current_value
            wb.save('commented_sample.xlsx')"""


#df1.set_index('Registrar', inplace=True)
#print(df1)

#print(df1.loc[~df1['Registrar'].str.contains('Delaware|aCRA', flags=re.I, regex=True, na=False)])#search for the inverse items based of on the inversed regex that contains the specific fields






















